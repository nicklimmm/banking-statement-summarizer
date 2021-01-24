from getpass import getpass
from openpyxl import Workbook
from openpyxl.styles import colors, Font, Color, Alignment, NamedStyle, PatternFill
from openpyxl.chart import Reference, ScatterChart, Series
from PyPDF2 import PdfFileReader
from pikepdf import Pdf, PasswordError

import re
import os


def extract_data_from_pdf(cwd=os.getcwd()):
    # when typing the password, it will not be shown in the terminal
    password = getpass('Enter password: ')
    temp_dir = cwd + '/temp.pdf'

    # to store the data based on the year (key = year, value = [month, inflow, outflow, netflow, interest, avg_balance])
    data_dict = dict()

    for file_name in os.listdir(cwd):
        # to search for FRANK OCBC e-Statements in the same folder
        if file_name.startswith("FRANK") and file_name.endswith(".pdf"):
            file_dir = cwd + '/' + file_name

            try:
                '''since PyPDF2 cannot open the encrypted file, we use pikepdf
                to open the file and create a copy in a temporary pdf file that
                is decrypted for extraction
                '''
                temp_file = Pdf.open(file_dir, password=password)
                temp_file.save(temp_dir)
                pdf_obj = PdfFileReader(temp_dir)
            except PasswordError:
                print(
                    'Wrong password! Please rerun the script again to create the summary.')
                exit()

            try:
                date_created = pdf_obj.getDocumentInfo()['/CreationDate']
                year = int(date_created[2:6])
                # the statement is created 1 month later, so we decrement by 1 to get the actual data
                month = int(date_created[6:8]) - 1

                # error handling when the e-statement is received on January (which is e-statement for December)
                if month == 0:
                    year -= 1
                    month = 12

                # to handle different number of pages in each file, and the summary lies in the back pages of the file
                num_pages = pdf_obj.getNumPages()
                if num_pages == 3:
                    page_obj = pdf_obj.getPage(num_pages - 3)
                else:
                    page_obj = pdf_obj.getPage(num_pages - 2)

                # using regex to find the necessary details and extract those to variables
                text = page_obj.extractText().encode('ascii').decode('ascii')
                pattern = r'[0-9,\.]+\s+[0-9,\.]+\s+[0-9,\.]+\s+[0-9,\.]+[0-9]'
                result = re.findall(pattern, text)
                inflow, outflow, interest, avg_balance = list(
                    map(float, result[-1].replace(',', '').split()))
                netflow = round(float(inflow) - float(outflow), 2)

                if year not in data_dict:
                    data_dict[year] = []
                data_dict[year].append((month, inflow, outflow, netflow))
            except:
                print('Something went wrong... Please rerun the script or report the issue in GitHub.')
                os.remove(temp_dir)
                exit()

    # to prevent unauthorized access the decrypted pdf
    os.remove(temp_dir)

    return data_dict


def create_annual_chart(worksheet=None, year=0, min_row=1, max_row=1):
    chart = ScatterChart()

    # sets the chart styling
    chart.title = f'{year}'
    chart.x_axis.title = 'Month'
    chart.y_axis.title = 'Amount'
    chart.legend.position = 'b'
    chart.height = 7.7
    chart.width = 21.5

    xvalues = Reference(
        worksheet=worksheet,
        min_col=2,
        min_row=min_row + 1,
        max_row=max_row
    )

    for col in range(3, 6):
        values = Reference(
            worksheet=worksheet,
            min_col=col,
            min_row=min_row,
            max_row=max_row
        )
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)

    return chart


def insert_data_to_excel(worksheet=None, data_dict=dict()):
    # styling for headers
    header = NamedStyle(name='header')
    header.font = Font(bold=True)
    header.alignment = Alignment(horizontal='center', vertical='center')

    # to keep track on which row should we write next (useful when getting the cell range for calculations)
    next_row = 1

    # sort everything based on the date and insert the data into the excel sheet
    for year in sorted(data_dict):
        data_dict[year].sort()

        worksheet.append(['YEAR', 'MONTH', 'INFLOW', 'OUTFLOW', 'NETFLOW'])

        for col in ('A', 'B', 'C', 'D', 'E'):
            worksheet[f'{col}{next_row}'].style = header

        next_row += 1

        # counts how many e-Statements in the same year
        num_of_months = len(data_dict[year])

        for month, *data in data_dict[year]:
            worksheet.append([year, month, *data])
            next_row += 1

        # row ranges for the actual data (month, inflow, outflow, netflow)
        start_range = next_row - num_of_months
        end_range = next_row - 1

        # appends the TOTAL and AVERAGE rows and its data into the worksheet
        for desc, op in (('TOTAL', 'SUM'), ('AVERAGE', 'AVERAGE')):
            worksheet.merge_cells(f'A{next_row}:B{next_row}')
            cell = worksheet[f'A{next_row}']
            cell.style = header
            cell.value = f'{desc} ({year})'
            for col in ('C', 'D', 'E'):
                cell = worksheet[f'{col}{next_row}']
                cell.value = f'={op}({col}{start_range}:{col}{end_range})'
            next_row += 1

        # create and add the chart into the worksheet
        chart = create_annual_chart(
            worksheet=worksheet,
            year=year,
            min_row=start_range - 1,
            max_row=end_range
        )
        worksheet.add_chart(chart, f'G{start_range - 1}')

        # padding for filling the missing months
        for row in range(12 - num_of_months):
            worksheet.append([])
            next_row += 1

        worksheet.append([])    # spacing for the next year
        worksheet.append([])
        next_row += 2


if __name__ == '__main__':
    cwd = os.getcwd()   # cwd = current working directory

    excel_dir = cwd + '/summary_excel.xlsx'

    data_dict = extract_data_from_pdf(cwd=cwd)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    insert_data_to_excel(worksheet=ws, data_dict=data_dict)

    wb.save(filename=excel_dir)
